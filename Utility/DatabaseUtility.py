import openpyxl
import pyodbc

import DBConfigurations.Config_file


class DatabaseConnection:
    def Query(self,row_num):

        cnxn_str = ("Driver={SQL Server};""Server=RLPT-0015\SQLEXPRESS;""Database=POC2;""Trusted_Connection=yes;")

        my_cnxn = pyodbc.connect(cnxn_str)

        cursor = my_cnxn.cursor()
        book = openpyxl.load_workbook(DBConfigurations.Config_file.excelInputQuerry)
        sheet = book.active
        cell = sheet.cell(row=row_num, column=3)

        query=str(cell.value)

        cursor.execute(query)

        db_list = []

        for i in cursor:
            db_list.append(i)

        return db_list

    def Query36(self,):

        cnxn_str = ("Driver={SQL Server};""Server=RLPT-0015\SQLEXPRESS;""Database=POC2;""Trusted_Connection=yes;")

        my_cnxn = pyodbc.connect(cnxn_str)

        cursor = my_cnxn.cursor()
        book = openpyxl.load_workbook(DBConfigurations.Config_file.excelInputQuerry)
        sheet = book.active
        cell = sheet.cell(row=36, column=3)

        query=str(cell.value)

        cursor.execute(query)

        db_list = []

        for i in cursor:
            db_list.append(i)

        return db_list




